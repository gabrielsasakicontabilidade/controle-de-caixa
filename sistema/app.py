# -*- coding: utf-8 -*-
import io
import os
import re
from datetime import date, datetime, timedelta
from functools import wraps

from flask import Flask, render_template, request, redirect, url_for, flash, send_file, session
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from werkzeug.security import generate_password_hash, check_password_hash

from models import (
    db, Config, MovimentoCaixa, ContaPagar, ContaReceber, DreMensal, Usuario,
    CATEGORIAS_MOVIMENTO, CATEGORIAS_ENTRADA, CATEGORIAS_PAGAR, CATEGORIAS_RECEBER,
    MESES_PT, hoje_brasil,
)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# DATABASE_URL is provided by the hosting platform (e.g. Neon/Vercel) in production.
# Falls back to a local SQLite file for local development.
_db_url = os.environ.get("DATABASE_URL")
if _db_url:
    if _db_url.startswith("postgres://"):
        _db_url = _db_url.replace("postgres://", "postgresql://", 1)
else:
    _db_url = "sqlite:///" + os.path.join(BASE_DIR, "caixa.db")

app = Flask(__name__)
app.config["SQLALCHEMY_DATABASE_URI"] = _db_url
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "supermercado-fluxo-caixa-dev-local")
db.init_app(app)

APP_USERNAME = os.environ.get("APP_USERNAME", "admin")
APP_PASSWORD = os.environ.get("APP_PASSWORD", "admin123")

DIAS_PT = ["Segunda-feira", "Terca-feira", "Quarta-feira", "Quinta-feira", "Sexta-feira", "Sabado", "Domingo"]
MESES_EXT_PT = ["", "janeiro", "fevereiro", "marco", "abril", "maio", "junho", "julho",
                "agosto", "setembro", "outubro", "novembro", "dezembro"]


@app.context_processor
def inject_hoje_extenso():
    hoje = hoje_brasil()
    return {"hoje_extenso": f"{DIAS_PT[hoje.weekday()]}, {hoje.day} de {MESES_EXT_PT[hoje.month]} de {hoje.year}"}


def login_required(view_fn):
    @wraps(view_fn)
    def wrapped(*args, **kwargs):
        if not session.get("logado"):
            return redirect(url_for("login", next=request.path))
        return view_fn(*args, **kwargs)
    return wrapped


def admin_required(view_fn):
    @wraps(view_fn)
    def wrapped(*args, **kwargs):
        if not session.get("logado"):
            return redirect(url_for("login", next=request.path))
        if not session.get("admin"):
            flash("Apenas administradores podem acessar esta pagina.", "danger")
            return redirect(url_for("dashboard"))
        return view_fn(*args, **kwargs)
    return wrapped


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        usuario = request.form.get("usuario", "")
        senha = request.form.get("senha", "")
        u = Usuario.query.filter_by(username=usuario, ativo=True).first()
        if u and check_password_hash(u.password_hash, senha):
            session["logado"] = True
            session["usuario"] = u.username
            session["admin"] = bool(u.admin)
            destino = request.args.get("next") or url_for("dashboard")
            return redirect(destino)
        flash("Usuario ou senha invalidos.", "danger")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.before_request
def exigir_login():
    rotas_livres = {"login", "static"}
    if request.endpoint and request.endpoint not in rotas_livres and not session.get("logado"):
        return redirect(url_for("login", next=request.path))


@app.template_filter("brl")
def brl(value):
    try:
        v = float(value)
    except (TypeError, ValueError):
        return value
    s = f"{v:,.2f}"
    s = s.replace(",", "X").replace(".", ",").replace("X", ".")
    sign = ""
    if s.startswith("-"):
        sign = "-"
        s = s[1:]
    return f"{sign}R$ {s}"


@app.template_filter("pct")
def pct(value):
    try:
        return f"{float(value) * 100:.1f}%"
    except (TypeError, ValueError):
        return value


@app.template_filter("brdate")
def brdate(value):
    if not value:
        return ""
    return value.strftime("%d/%m/%Y")


@app.template_filter("statusclass")
def statusclass(value):
    return (value or "").replace(" ", "")


@app.template_filter("badgeclass")
def badgeclass(value):
    if value in ("Pago", "Recebido"):
        return "badge-green"
    if value == "Atrasado":
        return "badge-red"
    return "badge-amber"


def parse_date(value):
    if not value:
        return None
    return datetime.strptime(value, "%Y-%m-%d").date()


def parse_valor(value, allow_negative=False):
    """Aceita '15000', '15000.50', '15000,50', '15.000,00', 'R$ 15.000,00', '2,5%' etc."""
    if value is None:
        raise ValueError("Valor vazio.")
    s = str(value).strip()
    s = re.sub(r"[Rr]\$\s*", "", s)
    s = s.replace("%", "")
    s = s.replace(" ", "")
    if not s:
        raise ValueError("Valor vazio.")
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    if not re.fullmatch(r"-?\d+(\.\d+)?", s):
        raise ValueError(f"Valor invalido: '{value}'.")
    resultado = float(s)
    if resultado < 0 and not allow_negative:
        raise ValueError(f"Valor nao pode ser negativo: '{value}'.")
    return resultado


def add_months(year, month, n):
    total = (year * 12 + (month - 1)) + n
    return total // 12, total % 12 + 1


def month_range(year, month):
    start = date(year, month, 1)
    ny, nm = add_months(year, month, 1)
    end = date(ny, nm, 1)
    return start, end


def get_config():
    cfg = Config.query.first()
    if not cfg:
        cfg = Config(saldo_inicial=0.0, data_inicial=hoje_brasil())
        db.session.add(cfg)
        db.session.commit()
    return cfg


def saldo_atual():
    cfg = get_config()
    total_entrada = sum(m.entrada for m in MovimentoCaixa.query.all())
    total_saida = sum(m.saida for m in MovimentoCaixa.query.all())
    return cfg.saldo_inicial + total_entrada - total_saida


# ---------------------------------------------------------------- Dashboard
@app.route("/")
def dashboard():
    hoje = hoje_brasil()
    pagar_aberto = ContaPagar.query.filter(ContaPagar.data_pagamento.is_(None)).all()
    receber_aberto = ContaReceber.query.filter(ContaReceber.data_recebimento.is_(None)).all()

    total_pagar_aberto = sum(c.valor for c in pagar_aberto)
    total_pagar_atrasado = sum(c.valor for c in pagar_aberto if c.status == "Atrasado")
    total_receber_aberto = sum(c.valor_liquido for c in receber_aberto)
    total_receber_atrasado = sum(c.valor_liquido for c in receber_aberto if c.status == "Atrasado")

    ultimos_mov = MovimentoCaixa.query.order_by(MovimentoCaixa.data.desc(), MovimentoCaixa.id.desc()).limit(8).all()
    vencendo = sorted(
        [c for c in pagar_aberto if 0 <= (c.data_vencimento - hoje).days <= 7],
        key=lambda c: c.data_vencimento,
    )

    return render_template(
        "dashboard.html",
        saldo=saldo_atual(),
        total_pagar_aberto=total_pagar_aberto,
        total_pagar_atrasado=total_pagar_atrasado,
        total_receber_aberto=total_receber_aberto,
        total_receber_atrasado=total_receber_atrasado,
        ultimos_mov=ultimos_mov,
        vencendo=vencendo,
        hoje=hoje,
    )


@app.route("/config", methods=["GET", "POST"])
@admin_required
def config_view():
    cfg = get_config()
    if request.method == "POST":
        try:
            cfg.saldo_inicial = parse_valor(request.form["saldo_inicial"])
        except ValueError as e:
            flash(str(e), "danger")
            return redirect(url_for("config_view"))
        cfg.data_inicial = parse_date(request.form["data_inicial"]) or cfg.data_inicial
        db.session.commit()
        flash("Configuracao salva.", "success")
        return redirect(url_for("dashboard"))
    return render_template("config.html", cfg=cfg)


# ---------------------------------------------------------- Movimento Caixa
@app.route("/movimento", methods=["GET", "POST"])
def movimento_list():
    if request.method == "POST":
        try:
            valor = parse_valor(request.form["valor"])
        except ValueError as e:
            flash(str(e), "danger")
            return redirect(url_for("movimento_list"))
        m = MovimentoCaixa(
            data=parse_date(request.form["data"]) or hoje_brasil(),
            descricao=request.form["descricao"],
            categoria=request.form["categoria"],
            valor=valor,
        )
        db.session.add(m)
        db.session.commit()
        flash("Lancamento adicionado.", "success")
        return redirect(url_for("movimento_list"))

    cfg = get_config()
    itens = MovimentoCaixa.query.order_by(MovimentoCaixa.data.asc(), MovimentoCaixa.id.asc()).all()
    saldo = cfg.saldo_inicial
    linhas = []
    total_entradas = 0.0
    total_saidas = 0.0
    for m in itens:
        total_entradas += m.entrada
        total_saidas += m.saida
        saldo += m.entrada - m.saida
        linhas.append((m, saldo))
    linhas.reverse()
    return render_template(
        "movimento_list.html",
        linhas=linhas,
        categorias=CATEGORIAS_MOVIMENTO,
        categorias_entrada=CATEGORIAS_ENTRADA,
        saldo_inicial=cfg.saldo_inicial,
        total_entradas=total_entradas,
        total_saidas=total_saidas,
        saldo_periodo=total_entradas - total_saidas,
        hoje=hoje_brasil().isoformat(),
    )


@app.route("/movimento/<int:item_id>/excluir", methods=["POST"])
def movimento_excluir(item_id):
    m = MovimentoCaixa.query.get_or_404(item_id)
    db.session.delete(m)
    db.session.commit()
    flash("Lancamento removido.", "success")
    return redirect(url_for("movimento_list"))


# ----------------------------------------------------------- Contas a Pagar
@app.route("/contas-a-pagar", methods=["GET", "POST"])
def pagar_list():
    if request.method == "POST":
        try:
            valor = parse_valor(request.form["valor"])
        except ValueError as e:
            flash(str(e), "danger")
            return redirect(url_for("pagar_list"))
        c = ContaPagar(
            fornecedor=request.form["fornecedor"],
            descricao=request.form.get("descricao"),
            categoria=request.form["categoria"],
            data_emissao=parse_date(request.form.get("data_emissao")),
            data_vencimento=parse_date(request.form["data_vencimento"]),
            valor=valor,
            observacoes=request.form.get("observacoes"),
        )
        db.session.add(c)
        db.session.commit()
        flash("Conta a pagar adicionada.", "success")
        return redirect(url_for("pagar_list"))

    itens = ContaPagar.query.order_by(ContaPagar.data_vencimento.asc()).all()
    total_aberto = sum(c.valor for c in itens if c.status == "Em Aberto")
    total_atrasado = sum(c.valor for c in itens if c.status == "Atrasado")
    total_pago = sum(c.valor for c in itens if c.status == "Pago")
    return render_template(
        "pagar_list.html", itens=itens, categorias=CATEGORIAS_PAGAR,
        total_aberto=total_aberto, total_atrasado=total_atrasado, total_pago=total_pago,
    )


CATEGORIA_MOVIMENTO_SAIDA_POR_PAGAR = {
    "Mercadorias/Fornecedores": "Pagamento Fornecedor",
}


@app.route("/contas-a-pagar/<int:item_id>/pagar", methods=["POST"])
def pagar_marcar_pago(item_id):
    c = ContaPagar.query.get_or_404(item_id)
    if c.data_pagamento:
        flash(f"Conta de {c.fornecedor} ja estava marcada como paga.", "danger")
        return redirect(url_for("pagar_list"))
    c.data_pagamento = hoje_brasil()
    categoria_mov = CATEGORIA_MOVIMENTO_SAIDA_POR_PAGAR.get(c.categoria, "Despesa Operacional")
    descricao = f"Pagamento: {c.fornecedor}" + (f" - {c.descricao}" if c.descricao else "")
    mov = MovimentoCaixa(data=c.data_pagamento, descricao=descricao, categoria=categoria_mov, valor=c.valor)
    db.session.add(mov)
    db.session.flush()
    c.movimento_id = mov.id
    db.session.commit()
    flash(f"Conta de {c.fornecedor} marcada como paga e lancada no Movimento de Caixa.", "success")
    return redirect(url_for("pagar_list"))


@app.route("/contas-a-pagar/<int:item_id>/excluir", methods=["POST"])
def pagar_excluir(item_id):
    c = ContaPagar.query.get_or_404(item_id)
    if c.movimento_id:
        mov = MovimentoCaixa.query.get(c.movimento_id)
        if mov:
            db.session.delete(mov)
    db.session.delete(c)
    db.session.commit()
    flash("Conta a pagar removida (e o lancamento de caixa vinculado, se havia).", "success")
    return redirect(url_for("pagar_list"))


# --------------------------------------------------------- Contas a Receber
@app.route("/contas-a-receber", methods=["GET", "POST"])
def receber_list():
    if request.method == "POST":
        try:
            valor_bruto = parse_valor(request.form["valor_bruto"])
            taxa_percentual = parse_valor(request.form.get("taxa_percentual") or "0") / 100.0
        except ValueError as e:
            flash(str(e), "danger")
            return redirect(url_for("receber_list"))
        c = ContaReceber(
            cliente=request.form["cliente"],
            descricao=request.form.get("descricao"),
            categoria=request.form["categoria"],
            data_venda=parse_date(request.form.get("data_venda")),
            data_prevista=parse_date(request.form["data_prevista"]),
            valor_bruto=valor_bruto,
            taxa_percentual=taxa_percentual,
            observacoes=request.form.get("observacoes"),
        )
        db.session.add(c)
        db.session.commit()
        flash("Conta a receber adicionada.", "success")
        return redirect(url_for("receber_list"))

    clientes_disponiveis = [
        row[0] for row in
        db.session.query(ContaReceber.cliente).distinct().order_by(ContaReceber.cliente.asc()).all()
    ]

    filtro_cliente = (request.args.get("cliente") or "").strip()
    query = ContaReceber.query
    if filtro_cliente:
        query = query.filter(ContaReceber.cliente.ilike(f"%{filtro_cliente}%"))
    itens = query.order_by(ContaReceber.data_prevista.asc()).all()

    total_aberto = sum(c.valor_liquido for c in itens if c.status == "Em Aberto")
    total_atrasado = sum(c.valor_liquido for c in itens if c.status == "Atrasado")
    total_recebido = sum(c.valor_liquido for c in itens if c.status == "Recebido")
    return render_template(
        "receber_list.html", itens=itens, categorias=CATEGORIAS_RECEBER,
        clientes_disponiveis=clientes_disponiveis, filtro_cliente=filtro_cliente,
        total_aberto=total_aberto, total_atrasado=total_atrasado, total_recebido=total_recebido,
    )


CATEGORIA_MOVIMENTO_ENTRADA_POR_RECEBER = {
    "Cartao de Credito": "Venda Cartao Credito",
    "Cartao de Debito": "Venda Cartao Debito",
}


@app.route("/contas-a-receber/<int:item_id>/receber", methods=["POST"])
def receber_marcar_recebido(item_id):
    c = ContaReceber.query.get_or_404(item_id)
    if c.data_recebimento:
        flash(f"Recebimento de {c.cliente} ja estava confirmado.", "danger")
        return redirect(url_for("receber_list"))
    c.data_recebimento = hoje_brasil()
    categoria_mov = CATEGORIA_MOVIMENTO_ENTRADA_POR_RECEBER.get(c.categoria, "Outras Entradas")
    descricao = f"Recebimento: {c.cliente}" + (f" - {c.descricao}" if c.descricao else "")
    mov = MovimentoCaixa(data=c.data_recebimento, descricao=descricao, categoria=categoria_mov, valor=c.valor_liquido)
    db.session.add(mov)
    db.session.flush()
    c.movimento_id = mov.id
    db.session.commit()
    flash(f"Recebimento de {c.cliente} confirmado e lancado no Movimento de Caixa.", "success")
    return redirect(url_for("receber_list"))


@app.route("/contas-a-receber/<int:item_id>/excluir", methods=["POST"])
def receber_excluir(item_id):
    c = ContaReceber.query.get_or_404(item_id)
    if c.movimento_id:
        mov = MovimentoCaixa.query.get(c.movimento_id)
        if mov:
            db.session.delete(mov)
    db.session.delete(c)
    db.session.commit()
    flash("Conta a receber removida (e o lancamento de caixa vinculado, se havia).", "success")
    return redirect(url_for("receber_list"))


# ----------------------------------------------------- Fluxo Caixa Projetado
@app.route("/fluxo-projetado")
def fluxo_projetado():
    try:
        otimista = parse_valor(request.args.get("otimista", 10), allow_negative=True) / 100.0
    except ValueError:
        otimista = 0.10
    try:
        pessimista = parse_valor(request.args.get("pessimista", -10), allow_negative=True) / 100.0
    except ValueError:
        pessimista = -0.10

    hoje = hoje_brasil()
    meses = []
    y, m = hoje.year, hoje.month
    for i in range(12):
        meses.append((y, m))
        y, m = add_months(y, m, 1)

    pagar_all = ContaPagar.query.all()
    receber_all = ContaReceber.query.all()

    colunas = []
    saldo_ini = saldo_atual()
    for (ano, mes) in meses:
        ini, fim = month_range(ano, mes)
        entradas_prev = sum(c.valor_liquido for c in receber_all if ini <= c.data_prevista < fim)
        saidas_prev = sum(c.valor for c in pagar_all if ini <= c.data_vencimento < fim)
        saldo_final = saldo_ini + entradas_prev - saidas_prev
        colunas.append({
            "label": f"{MESES_PT[mes]}/{str(ano)[2:]}",
            "saldo_inicial": saldo_ini,
            "entradas": entradas_prev,
            "saidas": saidas_prev,
            "saldo_final": saldo_final,
            "saldo_otimista": saldo_final + entradas_prev * otimista,
            "saldo_pessimista": saldo_final + entradas_prev * pessimista,
        })
        saldo_ini = saldo_final

    return render_template(
        "fluxo_projetado.html", colunas=colunas,
        otimista=int(otimista * 100), pessimista=int(pessimista * 100),
    )


# ------------------------------------------------------------- DRE Gerencial
@app.route("/dre")
def dre_list():
    ano = int(request.args.get("ano", hoje_brasil().year))
    linhas = []
    for mes in range(1, 13):
        d = DreMensal.query.filter_by(ano=ano, mes=mes).first()
        linhas.append((mes, d))
    return render_template("dre_list.html", ano=ano, linhas=linhas, meses_pt=MESES_PT)


@app.route("/dre/<int:ano>/<int:mes>/editar", methods=["GET", "POST"])
def dre_editar(ano, mes):
    d = DreMensal.query.filter_by(ano=ano, mes=mes).first()
    if not d:
        d = DreMensal(ano=ano, mes=mes)
        db.session.add(d)
        db.session.commit()

    if request.method == "POST":
        def f(name):
            return parse_valor(request.form.get(name) or "0")
        try:
            valores = {name: f(name) for name in [
                "receita_bruta", "devolucoes", "aliquota_impostos", "cmv", "salarios",
                "aluguel", "energia_agua", "materiais_embalagens", "marketing",
                "outras_despesas", "despesas_financeiras", "receitas_financeiras", "numero_vendas",
            ]}
        except ValueError as e:
            flash(str(e), "danger")
            return redirect(url_for("dre_editar", ano=ano, mes=mes))
        d.receita_bruta = valores["receita_bruta"]
        d.devolucoes = valores["devolucoes"]
        d.aliquota_impostos = valores["aliquota_impostos"] / 100.0
        d.cmv = valores["cmv"]
        d.salarios = valores["salarios"]
        d.aluguel = valores["aluguel"]
        d.energia_agua = valores["energia_agua"]
        d.materiais_embalagens = valores["materiais_embalagens"]
        d.marketing = valores["marketing"]
        d.outras_despesas = valores["outras_despesas"]
        d.despesas_financeiras = valores["despesas_financeiras"]
        d.receitas_financeiras = valores["receitas_financeiras"]
        d.numero_vendas = int(valores["numero_vendas"])
        db.session.commit()
        flash(f"DRE de {MESES_PT[mes]}/{ano} salva.", "success")
        return redirect(url_for("dre_list", ano=ano))

    return render_template("dre_form.html", d=d, ano=ano, mes=mes, meses_pt=MESES_PT)


# --------------------------------------------------------------------- Excel
@app.route("/exportar-excel")
def exportar_excel():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="2E7D32")
    header_font = Font(color="FFFFFF", bold=True)

    def add_sheet(name, headers, rows):
        ws = wb.create_sheet(name)
        for i, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = header_font
            c.fill = header_fill
        for r, row in enumerate(rows, start=2):
            for i, val in enumerate(row, start=1):
                ws.cell(row=r, column=i, value=val)
        for i in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 20

    cfg = get_config()
    mov = MovimentoCaixa.query.order_by(MovimentoCaixa.data.asc()).all()
    add_sheet("Movimento_Caixa_Diario",
              ["Data", "Descricao", "Categoria", "Tipo", "Entrada", "Saida"],
              [[m.data.strftime("%d/%m/%Y"), m.descricao, m.categoria, m.tipo, m.entrada, m.saida] for m in mov])

    pagar = ContaPagar.query.all()
    add_sheet("Contas_a_Pagar",
              ["Fornecedor", "Descricao", "Categoria", "Data Vencimento", "Valor", "Status", "Data Pagamento"],
              [[c.fornecedor, c.descricao, c.categoria, c.data_vencimento.strftime("%d/%m/%Y"),
                c.valor, c.status, c.data_pagamento.strftime("%d/%m/%Y") if c.data_pagamento else ""] for c in pagar])

    receber = ContaReceber.query.all()
    add_sheet("Contas_a_Receber",
              ["Cliente", "Descricao", "Categoria", "Data Prevista", "Valor Bruto", "Taxa", "Valor Liquido", "Status"],
              [[c.cliente, c.descricao, c.categoria, c.data_prevista.strftime("%d/%m/%Y"),
                c.valor_bruto, c.taxa_percentual, c.valor_liquido, c.status] for c in receber])

    dre_rows = DreMensal.query.order_by(DreMensal.ano.asc(), DreMensal.mes.asc()).all()
    add_sheet("DRE_Gerencial",
              ["Ano", "Mes", "Receita Bruta", "Receita Liquida", "Lucro Bruto", "Resultado Operacional", "Resultado Liquido"],
              [[d.ano, MESES_PT[d.mes], d.receita_bruta, d.receita_liquida, d.lucro_bruto,
                d.resultado_operacional, d.resultado_liquido] for d in dre_rows])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"Fluxo_Caixa_Exportado_{hoje_brasil().strftime('%Y%m%d')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ------------------------------------------------------------------- Usuarios
@app.route("/usuarios", methods=["GET", "POST"])
@admin_required
def usuarios_list():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        senha = request.form.get("senha", "")
        is_admin = request.form.get("admin") == "on"
        if not username or not senha:
            flash("Usuario e senha sao obrigatorios.", "danger")
        elif Usuario.query.filter_by(username=username).first():
            flash("Ja existe um usuario com esse nome.", "danger")
        else:
            u = Usuario(username=username, password_hash=generate_password_hash(senha), ativo=True, admin=is_admin)
            db.session.add(u)
            db.session.commit()
            flash(f"Usuario '{username}' criado.", "success")
        return redirect(url_for("usuarios_list"))

    itens = Usuario.query.order_by(Usuario.username.asc()).all()
    return render_template("usuarios_list.html", itens=itens)


@app.route("/usuarios/<int:item_id>/excluir", methods=["POST"])
@admin_required
def usuarios_excluir(item_id):
    u = Usuario.query.get_or_404(item_id)
    if u.username == session.get("usuario"):
        flash("Voce nao pode excluir o proprio usuario logado.", "danger")
        return redirect(url_for("usuarios_list"))
    db.session.delete(u)
    db.session.commit()
    flash(f"Usuario '{u.username}' removido.", "success")
    return redirect(url_for("usuarios_list"))


@app.route("/usuarios/<int:item_id>/redefinir-senha", methods=["POST"])
@admin_required
def usuarios_redefinir_senha(item_id):
    u = Usuario.query.get_or_404(item_id)
    nova_senha = request.form.get("nova_senha", "")
    if not nova_senha:
        flash("Informe uma nova senha.", "danger")
    else:
        u.password_hash = generate_password_hash(nova_senha)
        db.session.commit()
        flash(f"Senha de '{u.username}' redefinida.", "success")
    return redirect(url_for("usuarios_list"))


def _seed_admin_inicial():
    if Usuario.query.count() == 0:
        db.session.add(Usuario(
            username=APP_USERNAME,
            password_hash=generate_password_hash(APP_PASSWORD),
            ativo=True,
            admin=True,
        ))
        db.session.commit()


def _promover_admin_inicial():
    """Garante que exista pelo menos um administrador. Cobre a migracao: usuarios
    criados antes da coluna 'admin' existir ficam com admin=False por padrao."""
    if Usuario.query.filter_by(admin=True).count() == 0:
        primeiro = Usuario.query.order_by(Usuario.id.asc()).first()
        if primeiro:
            primeiro.admin = True
            db.session.commit()


def _ensure_column(table_name, column_name, column_type_sql):
    """db.create_all() so cria tabelas novas, nao adiciona colunas em tabelas
    ja existentes. Isso cobre a migracao leve das colunas adicionadas depois."""
    from sqlalchemy import inspect, text
    inspector = inspect(db.engine)
    if table_name not in inspector.get_table_names():
        return
    colunas = [c["name"] for c in inspector.get_columns(table_name)]
    if column_name not in colunas:
        with db.engine.begin() as conn:
            conn.execute(text(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {column_type_sql}"))


with app.app_context():
    db.create_all()
    _ensure_column("conta_pagar", "movimento_id", "INTEGER")
    _ensure_column("conta_receber", "movimento_id", "INTEGER")
    _ensure_column("usuario", "admin", "BOOLEAN DEFAULT FALSE")
    _seed_admin_inicial()
    _promover_admin_inicial()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8000, debug=False)
